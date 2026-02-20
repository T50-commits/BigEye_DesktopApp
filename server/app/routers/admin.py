"""
BigEye Pro — Admin API Router
All endpoints require admin authentication via require_admin dependency.

Sections:
  1. Dashboard Stats & Charts
  2. Users CRUD + Actions
  3. Slips Management
  4. Jobs Management
  5. Finance (Daily, Monthly, Export)
  6. System Config
  7. Audit Logs
"""
import logging
import math
from datetime import datetime, timezone, timedelta
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from google.cloud import firestore
from google.cloud.firestore_v1 import FieldFilter
from pydantic import BaseModel

from app.database import (
    users_ref, jobs_ref, transactions_ref, slips_ref,
    system_config_ref, audit_logs_ref, daily_reports_ref, get_db,
)
from app.dependencies import get_current_user
from app.config import settings
from app.security import hash_password, verify_password, create_jwt_token

logger = logging.getLogger("bigeye-api")
router = APIRouter(prefix="/admin", tags=["Admin"])

TH_TZ = timezone(timedelta(hours=7))


# ══════════════════════════════════════════════════
# Admin Dependency
# ══════════════════════════════════════════════════

async def require_admin(user: dict = Depends(get_current_user)) -> dict:
    """Verify that the current user is an admin."""
    user_id = user.get("user_id", "")
    if user_id not in settings.admin_uid_list:
        raise HTTPException(status_code=403, detail="Admin access required")
    return user


# ══════════════════════════════════════════════════
# Admin Login (no hardware_id required)
# ══════════════════════════════════════════════════

class AdminLoginRequest(BaseModel):
    email: str
    password: str


@router.post("/login")
async def admin_login(req: AdminLoginRequest):
    """Admin login — email + password only, no hardware_id check."""
    now = datetime.now(timezone.utc)

    docs = list(
        users_ref()
        .where(filter=FieldFilter("email", "==", req.email.lower()))
        .limit(1)
        .stream()
    )
    if not docs:
        raise HTTPException(status_code=401, detail="Invalid email or password")

    doc = docs[0]
    user = doc.to_dict()
    user_id = doc.id

    if not verify_password(req.password, user.get("password_hash", "")):
        raise HTTPException(status_code=401, detail="Invalid email or password")

    if user.get("status") in ("banned", "suspended"):
        raise HTTPException(status_code=403, detail="Account suspended")

    if user_id not in settings.admin_uid_list:
        raise HTTPException(status_code=403, detail="Admin access required")

    token = create_jwt_token(user_id, req.email.lower())

    audit_logs_ref().add({
        "event_type": "ADMIN_LOGIN",
        "user_id": user_id,
        "details": {"email": req.email.lower()},
        "severity": "INFO",
        "created_at": now,
    })

    logger.info(f"Admin login: {req.email}")
    return {
        "token": token,
        "user_id": user_id,
        "email": user.get("email", ""),
        "full_name": user.get("full_name", ""),
    }


# ══════════════════════════════════════════════════
# Helpers
# ══════════════════════════════════════════════════

def _ts_to_str(val) -> str:
    """Convert a Firestore timestamp to ISO string."""
    if val is None:
        return ""
    if hasattr(val, "isoformat"):
        return val.astimezone(TH_TZ).isoformat()
    return str(val)


def _today_range_utc():
    """Return (start_of_today_utc, end_of_today_utc) based on Thailand time."""
    now_th = datetime.now(TH_TZ)
    start_th = now_th.replace(hour=0, minute=0, second=0, microsecond=0)
    end_th = start_th + timedelta(days=1)
    return start_th.astimezone(timezone.utc), end_th.astimezone(timezone.utc)


def _paginate(items: list, page: int, limit: int) -> tuple[list, int]:
    """Return (page_items, total_count)."""
    total = len(items)
    start = (page - 1) * limit
    return items[start:start + limit], total


# ══════════════════════════════════════════════════
# 1. Dashboard Stats & Charts
# ══════════════════════════════════════════════════

@router.get("/dashboard/stats")
async def dashboard_stats(admin: dict = Depends(require_admin)):
    """Get dashboard summary statistics."""
    today_start, today_end = _today_range_utc()

    # Active users (status=active)
    all_users = list(users_ref().stream())
    active_users = sum(1 for u in all_users if u.to_dict().get("status") == "active")
    new_today = sum(
        1 for u in all_users
        if u.to_dict().get("created_at") and
        today_start <= u.to_dict()["created_at"].astimezone(timezone.utc) < today_end
    )

    # Topup today (verified slips)
    all_slips = list(slips_ref().where("status", "==", "VERIFIED").stream())
    topup_thb_today = 0.0
    for s in all_slips:
        sd = s.to_dict()
        created = sd.get("verified_at") or sd.get("created_at")
        if created and today_start <= created.astimezone(timezone.utc) < today_end:
            topup_thb_today += float(sd.get("amount_detected", 0) or 0)

    # Jobs today
    all_jobs = list(jobs_ref().stream())
    jobs_today = 0
    errors_today = 0
    recognized_credits_today = 0
    for j in all_jobs:
        jd = j.to_dict()
        created = jd.get("created_at")
        if created and today_start <= created.astimezone(timezone.utc) < today_end:
            jobs_today += 1
            if jd.get("status") == "COMPLETED":
                recognized_credits_today += jd.get("actual_usage", 0)
            if jd.get("failed_count", 0) > 0:
                errors_today += jd.get("failed_count", 0)

    # Exchange rate from config
    cfg_doc = system_config_ref().document("app_settings").get()
    cfg = cfg_doc.to_dict() if cfg_doc.exists else {}
    exchange_rate = cfg.get("exchange_rate", settings.EXCHANGE_RATE)
    recognized_thb_today = recognized_credits_today / exchange_rate if exchange_rate else 0

    success_rate = 0.0
    if jobs_today > 0:
        completed_today = sum(
            1 for j in all_jobs
            if j.to_dict().get("status") == "COMPLETED" and
            j.to_dict().get("created_at") and
            today_start <= j.to_dict()["created_at"].astimezone(timezone.utc) < today_end
        )
        success_rate = round(completed_today / jobs_today * 100, 1)

    # Pending slips
    pending_slips = sum(1 for s in slips_ref().where("status", "==", "PENDING").stream())

    # Stuck jobs (RESERVED for > 2 hours)
    cutoff = datetime.now(timezone.utc) - timedelta(hours=settings.JOB_EXPIRE_HOURS)
    stuck_jobs = sum(
        1 for j in all_jobs
        if j.to_dict().get("status") == "RESERVED" and
        j.to_dict().get("created_at") and
        j.to_dict()["created_at"].astimezone(timezone.utc) < cutoff
    )

    return {
        "active_users": active_users,
        "new_users_today": new_today,
        "topup_thb_today": round(topup_thb_today, 2),
        "recognized_thb_today": round(recognized_thb_today, 2),
        "exchange_rate": exchange_rate,
        "jobs_today": jobs_today,
        "errors_today": errors_today,
        "success_rate": success_rate,
        "pending_slips": pending_slips,
        "stuck_jobs": stuck_jobs,
    }


@router.get("/dashboard/charts")
async def dashboard_charts(
    days: int = Query(default=30, le=90),
    admin: dict = Depends(require_admin),
):
    """Get chart data for revenue and user growth over N days."""
    now_th = datetime.now(TH_TZ)
    start_date = (now_th - timedelta(days=days)).replace(hour=0, minute=0, second=0, microsecond=0)

    # Pre-fetch all data
    all_slips = [s.to_dict() for s in slips_ref().where("status", "==", "VERIFIED").stream()]
    all_jobs = [j.to_dict() for j in jobs_ref().where("status", "==", "COMPLETED").stream()]
    all_users = [u.to_dict() for u in users_ref().stream()]

    cfg_doc = system_config_ref().document("app_settings").get()
    cfg = cfg_doc.to_dict() if cfg_doc.exists else {}
    exchange_rate = cfg.get("exchange_rate", settings.EXCHANGE_RATE)

    revenue_data = []
    users_data = []

    for i in range(days):
        day = start_date + timedelta(days=i)
        day_end = day + timedelta(days=1)
        day_str = day.strftime("%Y-%m-%d")
        day_start_utc = day.astimezone(timezone.utc)
        day_end_utc = day_end.astimezone(timezone.utc)

        # Revenue
        topup = sum(
            float(s.get("amount_detected", 0) or 0)
            for s in all_slips
            if s.get("verified_at") and
            day_start_utc <= s["verified_at"].astimezone(timezone.utc) < day_end_utc
        )
        recognized_cr = sum(
            j.get("actual_usage", 0)
            for j in all_jobs
            if j.get("completed_at") and
            day_start_utc <= j["completed_at"].astimezone(timezone.utc) < day_end_utc
        )
        revenue_data.append({
            "date": day_str,
            "topup_thb": round(topup, 2),
            "recognized_thb": round(recognized_cr / exchange_rate, 2) if exchange_rate else 0,
        })

        # Users
        new_users = sum(
            1 for u in all_users
            if u.get("created_at") and
            day_start_utc <= u["created_at"].astimezone(timezone.utc) < day_end_utc
        )
        active = sum(
            1 for u in all_users
            if u.get("last_active") and
            day_start_utc <= u["last_active"].astimezone(timezone.utc) < day_end_utc
        )
        users_data.append({
            "date": day_str,
            "new_users": new_users,
            "active_users": active,
        })

    return {"revenue": revenue_data, "users": users_data}


# ══════════════════════════════════════════════════
# 2. Users
# ══════════════════════════════════════════════════

@router.get("/users")
async def list_users(
    search: str = Query(default="", max_length=200),
    page: int = Query(default=1, ge=1),
    limit: int = Query(default=50, ge=1, le=200),
    admin: dict = Depends(require_admin),
):
    """List users with optional search and pagination."""
    all_docs = list(users_ref().stream())
    items = []
    for doc in all_docs:
        d = doc.to_dict()
        d["uid"] = doc.id
        items.append(d)

    # Search filter
    if search:
        q = search.lower()
        items = [
            u for u in items
            if q in u.get("email", "").lower()
            or q in u.get("full_name", "").lower()
            or q in u.get("uid", "").lower()
        ]

    # Sort by created_at desc
    items.sort(
        key=lambda u: u.get("created_at") or datetime.min.replace(tzinfo=timezone.utc),
        reverse=True,
    )

    page_items, total = _paginate(items, page, limit)

    return {
        "users": [
            {
                "uid": u["uid"],
                "email": u.get("email", ""),
                "full_name": u.get("full_name", ""),
                "credits": u.get("credits", 0),
                "status": u.get("status", ""),
                "tier": u.get("tier", "standard"),
                "last_login": _ts_to_str(u.get("last_login")),
                "created_at": _ts_to_str(u.get("created_at")),
            }
            for u in page_items
        ],
        "total": total,
        "page": page,
        "pages": math.ceil(total / limit) if total else 1,
    }


@router.get("/users/{uid}")
async def get_user(uid: str, admin: dict = Depends(require_admin)):
    """Get detailed user info."""
    doc = users_ref().document(uid).get()
    if not doc.exists:
        raise HTTPException(status_code=404, detail="User not found")
    d = doc.to_dict()
    return {
        "uid": uid,
        "email": d.get("email", ""),
        "full_name": d.get("full_name", ""),
        "phone": d.get("phone", ""),
        "hardware_id": d.get("hardware_id", ""),
        "tier": d.get("tier", "standard"),
        "credits": d.get("credits", 0),
        "status": d.get("status", ""),
        "total_topup_baht": d.get("total_topup_baht", 0),
        "total_credits_used": d.get("total_credits_used", 0),
        "app_version": d.get("app_version", ""),
        "os_type": d.get("metadata", {}).get("os", ""),
        "created_at": _ts_to_str(d.get("created_at")),
        "last_login": _ts_to_str(d.get("last_login")),
        "last_active": _ts_to_str(d.get("last_active")),
    }


@router.get("/users/{uid}/transactions")
async def get_user_transactions(
    uid: str,
    limit: int = Query(default=50, le=200),
    admin: dict = Depends(require_admin),
):
    """Get user's credit transaction history."""
    docs = list(transactions_ref().where("user_id", "==", uid).stream())
    docs.sort(
        key=lambda d: d.to_dict().get("created_at") or datetime.min.replace(tzinfo=timezone.utc),
        reverse=True,
    )
    docs = docs[:limit]

    return {
        "transactions": [
            {
                "id": doc.id,
                "type": doc.to_dict().get("type", ""),
                "amount": doc.to_dict().get("amount", 0),
                "balance_after": doc.to_dict().get("balance_after", 0),
                "description": doc.to_dict().get("description", ""),
                "date": _ts_to_str(doc.to_dict().get("created_at")),
            }
            for doc in docs
        ]
    }


@router.get("/users/{uid}/jobs")
async def get_user_jobs(
    uid: str,
    limit: int = Query(default=50, le=200),
    admin: dict = Depends(require_admin),
):
    """Get user's job history."""
    docs = list(jobs_ref().where("user_id", "==", uid).stream())
    docs.sort(
        key=lambda d: d.to_dict().get("created_at") or datetime.min.replace(tzinfo=timezone.utc),
        reverse=True,
    )
    docs = docs[:limit]

    return {
        "jobs": [
            {
                "id": doc.id,
                "job_token": doc.to_dict().get("job_token", ""),
                "mode": doc.to_dict().get("mode", ""),
                "file_count": doc.to_dict().get("file_count", 0),
                "status": doc.to_dict().get("status", ""),
                "reserved_credits": doc.to_dict().get("reserved_credits", 0),
                "actual_usage": doc.to_dict().get("actual_usage", 0),
                "refund_amount": doc.to_dict().get("refund_amount", 0),
                "success_count": doc.to_dict().get("success_count", 0),
                "failed_count": doc.to_dict().get("failed_count", 0),
                "created_at": _ts_to_str(doc.to_dict().get("created_at")),
                "completed_at": _ts_to_str(doc.to_dict().get("completed_at")),
            }
            for doc in docs
        ]
    }


from pydantic import BaseModel, Field


class AdjustCreditsRequest(BaseModel):
    amount: int
    reason: str = ""


class ResetPasswordRequest(BaseModel):
    new_password: str = Field(min_length=8, max_length=128)
    reset_hardware: bool = False


@router.post("/users/{uid}/adjust-credits")
async def adjust_credits(uid: str, req: AdjustCreditsRequest, admin: dict = Depends(require_admin)):
    """Manually adjust a user's credits (positive or negative)."""
    user_doc = users_ref().document(uid).get()
    if not user_doc.exists:
        raise HTTPException(status_code=404, detail="User not found")

    now = datetime.now(timezone.utc)
    current = user_doc.to_dict().get("credits", 0)
    new_balance = current + req.amount

    users_ref().document(uid).update({"credits": new_balance})

    # Create transaction record
    transactions_ref().add({
        "user_id": uid,
        "type": "ADMIN_ADJUST",
        "amount": req.amount,
        "balance_after": new_balance,
        "description": f"ปรับเครดิตโดยแอดมิน: {req.reason}" if req.reason else "ปรับเครดิตโดยแอดมิน",
        "created_at": now,
    })

    # Audit
    audit_logs_ref().add({
        "event_type": "ADMIN_ADJUST_CREDITS",
        "user_id": admin["user_id"],
        "details": {"target_uid": uid, "amount": req.amount, "reason": req.reason, "new_balance": new_balance},
        "severity": "WARNING",
        "created_at": now,
    })

    return {"credits": new_balance, "message": f"ปรับเครดิต {req.amount:+d} เรียบร้อย"}


@router.post("/users/{uid}/suspend")
async def suspend_user(uid: str, admin: dict = Depends(require_admin)):
    """Suspend a user account."""
    user_doc = users_ref().document(uid).get()
    if not user_doc.exists:
        raise HTTPException(status_code=404, detail="User not found")
    if user_doc.to_dict().get("status") == "suspended":
        raise HTTPException(status_code=400, detail="User already suspended")

    now = datetime.now(timezone.utc)
    users_ref().document(uid).update({"status": "suspended"})

    audit_logs_ref().add({
        "event_type": "ADMIN_SUSPEND_USER",
        "user_id": admin["user_id"],
        "details": {"target_uid": uid},
        "severity": "WARNING",
        "created_at": now,
    })
    return {"message": "ระงับบัญชีเรียบร้อย"}


@router.post("/users/{uid}/unsuspend")
async def unsuspend_user(uid: str, admin: dict = Depends(require_admin)):
    """Reactivate a suspended user account."""
    user_doc = users_ref().document(uid).get()
    if not user_doc.exists:
        raise HTTPException(status_code=404, detail="User not found")
    if user_doc.to_dict().get("status") != "suspended":
        raise HTTPException(status_code=400, detail="User is not suspended")

    now = datetime.now(timezone.utc)
    users_ref().document(uid).update({"status": "active"})

    audit_logs_ref().add({
        "event_type": "ADMIN_UNSUSPEND_USER",
        "user_id": admin["user_id"],
        "details": {"target_uid": uid},
        "severity": "INFO",
        "created_at": now,
    })
    return {"message": "เปิดบัญชีเรียบร้อย"}


@router.post("/users/{uid}/reset-hardware")
async def reset_hardware(uid: str, admin: dict = Depends(require_admin)):
    """Reset a user's hardware ID so they can login from a new device."""
    user_doc = users_ref().document(uid).get()
    if not user_doc.exists:
        raise HTTPException(status_code=404, detail="User not found")

    now = datetime.now(timezone.utc)
    users_ref().document(uid).update({"hardware_id": ""})

    audit_logs_ref().add({
        "event_type": "ADMIN_RESET_HARDWARE",
        "user_id": admin["user_id"],
        "details": {"target_uid": uid},
        "severity": "WARNING",
        "created_at": now,
    })
    return {"message": "รีเซ็ต Hardware ID เรียบร้อย"}


@router.delete("/users/{uid}")
async def delete_user(uid: str, admin: dict = Depends(require_admin)):
    """Delete a user account and all associated data."""
    user_doc = users_ref().document(uid).get()
    if not user_doc.exists:
        raise HTTPException(status_code=404, detail="User not found")

    user_data = user_doc.to_dict()
    now = datetime.now(timezone.utc)

    # Safety: block deletion if user has remaining credits > 0
    if user_data.get("credits", 0) > 0:
        raise HTTPException(
            status_code=400,
            detail=f"ไม่สามารถลบบัญชีที่ยังมีเครดิตเหลืออยู่ ({user_data['credits']} เครดิต) — ปรับเครดิตเป็น 0 ก่อน"
        )

    users_ref().document(uid).delete()

    audit_logs_ref().add({
        "event_type": "ADMIN_DELETE_USER",
        "user_id": admin["user_id"],
        "details": {
            "target_uid": uid,
            "target_email": user_data.get("email", ""),
            "target_name": user_data.get("full_name", ""),
        },
        "severity": "CRITICAL",
        "created_at": now,
    })
    return {"message": f"ลบบัญชี {user_data.get('email', uid)} เรียบร้อย"}


@router.post("/users/{uid}/reset-password")
async def reset_password(uid: str, req: ResetPasswordRequest, admin: dict = Depends(require_admin)):
    """Reset a user's password."""
    user_doc = users_ref().document(uid).get()
    if not user_doc.exists:
        raise HTTPException(status_code=404, detail="User not found")

    now = datetime.now(timezone.utc)
    updates = {"password_hash": hash_password(req.new_password)}
    if req.reset_hardware:
        updates["hardware_id"] = ""

    users_ref().document(uid).update(updates)

    audit_logs_ref().add({
        "event_type": "ADMIN_RESET_PASSWORD",
        "user_id": admin["user_id"],
        "details": {"target_uid": uid, "reset_hardware": req.reset_hardware},
        "severity": "WARNING",
        "created_at": now,
    })
    return {"message": "รีเซ็ตรหัสผ่านเรียบร้อย"}


# ══════════════════════════════════════════════════
# 3. Slips
# ══════════════════════════════════════════════════

@router.get("/slips")
async def list_slips(
    status: str = Query(default="", max_length=20),
    page: int = Query(default=1, ge=1),
    limit: int = Query(default=50, ge=1, le=200),
    admin: dict = Depends(require_admin),
):
    """List slips with optional status filter."""
    if status:
        docs = list(slips_ref().where("status", "==", status.upper()).stream())
    else:
        docs = list(slips_ref().stream())

    items = []
    for doc in docs:
        d = doc.to_dict()
        d["id"] = doc.id
        items.append(d)

    items.sort(
        key=lambda s: s.get("created_at") or datetime.min.replace(tzinfo=timezone.utc),
        reverse=True,
    )

    page_items, total = _paginate(items, page, limit)

    return {
        "slips": [
            {
                "id": s["id"],
                "user_id": s.get("user_id", ""),
                "status": s.get("status", ""),
                "amount_detected": s.get("amount_detected"),
                "amount_credited": s.get("amount_credited"),
                "bank_ref": s.get("bank_ref", ""),
                "verification_method": s.get("verification_method", ""),
                "reject_reason": s.get("reject_reason", ""),
                "created_at": _ts_to_str(s.get("created_at")),
                "verified_at": _ts_to_str(s.get("verified_at")),
            }
            for s in page_items
        ],
        "total": total,
        "page": page,
        "pages": math.ceil(total / limit) if total else 1,
    }


@router.get("/slips/{slip_id}")
async def get_slip(slip_id: str, admin: dict = Depends(require_admin)):
    """Get detailed slip info."""
    doc = slips_ref().document(slip_id).get()
    if not doc.exists:
        raise HTTPException(status_code=404, detail="Slip not found")
    d = doc.to_dict()
    return {
        "id": slip_id,
        "user_id": d.get("user_id", ""),
        "status": d.get("status", ""),
        "amount_detected": d.get("amount_detected"),
        "amount_credited": d.get("amount_credited"),
        "bank_ref": d.get("bank_ref", ""),
        "verification_method": d.get("verification_method", ""),
        "verification_result": d.get("verification_result", {}),
        "reject_reason": d.get("reject_reason", ""),
        "metadata": d.get("metadata", {}),
        "created_at": _ts_to_str(d.get("created_at")),
        "verified_at": _ts_to_str(d.get("verified_at")),
    }


class ApproveSlipRequest(BaseModel):
    credit_amount: int = Field(gt=0)


class RejectSlipRequest(BaseModel):
    reason: str = ""


@router.post("/slips/{slip_id}/approve")
async def approve_slip(slip_id: str, req: ApproveSlipRequest, admin: dict = Depends(require_admin)):
    """Manually approve a pending slip and credit the user."""
    doc = slips_ref().document(slip_id).get()
    if not doc.exists:
        raise HTTPException(status_code=404, detail="Slip not found")
    d = doc.to_dict()
    if d.get("status") != "PENDING":
        raise HTTPException(status_code=400, detail=f"Slip is {d.get('status')}, not PENDING")

    now = datetime.now(timezone.utc)
    user_id = d.get("user_id", "")

    # Update slip
    slips_ref().document(slip_id).update({
        "status": "VERIFIED",
        "amount_credited": req.credit_amount,
        "verified_at": now,
        "verification_method": "MANUAL_ADMIN",
    })

    # Credit user
    user_doc = users_ref().document(user_id).get()
    if user_doc.exists:
        current = user_doc.to_dict().get("credits", 0)
        new_balance = current + req.credit_amount
        users_ref().document(user_id).update({"credits": new_balance})

        transactions_ref().add({
            "user_id": user_id,
            "type": "TOPUP",
            "amount": req.credit_amount,
            "balance_after": new_balance,
            "description": f"เติมเงิน (อนุมัติโดยแอดมิน) — {req.credit_amount} เครดิต",
            "created_at": now,
        })

    audit_logs_ref().add({
        "event_type": "ADMIN_APPROVE_SLIP",
        "user_id": admin["user_id"],
        "details": {"slip_id": slip_id, "target_uid": user_id, "credit_amount": req.credit_amount},
        "severity": "INFO",
        "created_at": now,
    })
    return {"message": f"อนุมัติสลิปเรียบร้อย — เพิ่ม {req.credit_amount} เครดิต"}


@router.post("/slips/{slip_id}/reject")
async def reject_slip(slip_id: str, req: RejectSlipRequest, admin: dict = Depends(require_admin)):
    """Reject a pending slip."""
    doc = slips_ref().document(slip_id).get()
    if not doc.exists:
        raise HTTPException(status_code=404, detail="Slip not found")
    d = doc.to_dict()
    if d.get("status") != "PENDING":
        raise HTTPException(status_code=400, detail=f"Slip is {d.get('status')}, not PENDING")

    now = datetime.now(timezone.utc)
    slips_ref().document(slip_id).update({
        "status": "REJECTED",
        "reject_reason": req.reason or "ปฏิเสธโดยแอดมิน",
    })

    audit_logs_ref().add({
        "event_type": "ADMIN_REJECT_SLIP",
        "user_id": admin["user_id"],
        "details": {"slip_id": slip_id, "target_uid": d.get("user_id"), "reason": req.reason},
        "severity": "INFO",
        "created_at": now,
    })
    return {"message": "ปฏิเสธสลิปเรียบร้อย"}


# ══════════════════════════════════════════════════
# 4. Jobs
# ══════════════════════════════════════════════════

@router.get("/jobs")
async def list_jobs(
    status: str = Query(default="", max_length=20),
    page: int = Query(default=1, ge=1),
    limit: int = Query(default=50, ge=1, le=200),
    admin: dict = Depends(require_admin),
):
    """List jobs with optional status filter."""
    if status:
        docs = list(jobs_ref().where("status", "==", status.upper()).stream())
    else:
        docs = list(jobs_ref().stream())

    items = []
    for doc in docs:
        d = doc.to_dict()
        d["id"] = doc.id
        items.append(d)

    items.sort(
        key=lambda j: j.get("created_at") or datetime.min.replace(tzinfo=timezone.utc),
        reverse=True,
    )

    page_items, total = _paginate(items, page, limit)

    # Resolve user names/emails for page items
    uid_set = {j.get("user_id", "") for j in page_items if j.get("user_id")}
    user_cache = {}
    for uid in uid_set:
        try:
            udoc = users_ref().document(uid).get()
            if udoc.exists:
                ud = udoc.to_dict()
                user_cache[uid] = {
                    "name": ud.get("full_name", ud.get("name", "")),
                    "email": ud.get("email", ""),
                }
        except Exception:
            pass

    return {
        "jobs": [
            {
                "id": j["id"],
                "job_token": j.get("job_token", ""),
                "user_id": j.get("user_id", ""),
                "user_name": user_cache.get(j.get("user_id", ""), {}).get("name", ""),
                "user_email": user_cache.get(j.get("user_id", ""), {}).get("email", ""),
                "mode": j.get("mode", ""),
                "file_count": j.get("file_count", 0),
                "status": j.get("status", ""),
                "reserved_credits": j.get("reserved_credits", 0),
                "actual_usage": j.get("actual_usage", 0),
                "success_count": j.get("success_count", 0),
                "failed_count": j.get("failed_count", 0),
                "created_at": _ts_to_str(j.get("created_at")),
            }
            for j in page_items
        ],
        "total": total,
        "page": page,
        "pages": math.ceil(total / limit) if total else 1,
    }


@router.get("/jobs/{job_id}")
async def get_job(job_id: str, admin: dict = Depends(require_admin)):
    """Get detailed job info."""
    doc = jobs_ref().document(job_id).get()
    if not doc.exists:
        raise HTTPException(status_code=404, detail="Job not found")
    d = doc.to_dict()
    return {
        "id": job_id,
        "job_token": d.get("job_token", ""),
        "user_id": d.get("user_id", ""),
        "mode": d.get("mode", ""),
        "keyword_style": d.get("keyword_style", ""),
        "model": d.get("model", ""),
        "file_count": d.get("file_count", 0),
        "photo_count": d.get("photo_count", 0),
        "video_count": d.get("video_count", 0),
        "status": d.get("status", ""),
        "reserved_credits": d.get("reserved_credits", 0),
        "actual_usage": d.get("actual_usage", 0),
        "refund_amount": d.get("refund_amount", 0),
        "photo_rate": d.get("photo_rate", 0),
        "video_rate": d.get("video_rate", 0),
        "success_count": d.get("success_count", 0),
        "failed_count": d.get("failed_count", 0),
        "version": d.get("version", ""),
        "created_at": _ts_to_str(d.get("created_at")),
        "completed_at": _ts_to_str(d.get("completed_at")),
    }


@router.post("/jobs/{job_id}/force-refund")
async def force_refund_job(job_id: str, admin: dict = Depends(require_admin)):
    """Force-refund a stuck/reserved job."""
    doc = jobs_ref().document(job_id).get()
    if not doc.exists:
        raise HTTPException(status_code=404, detail="Job not found")
    d = doc.to_dict()
    if d.get("status") not in ("RESERVED", "PROCESSING"):
        raise HTTPException(status_code=400, detail=f"Cannot refund job with status {d.get('status')}")

    now = datetime.now(timezone.utc)
    user_id = d.get("user_id", "")
    reserved = d.get("reserved_credits", 0)

    # Refund credits
    user_doc = users_ref().document(user_id).get()
    if user_doc.exists:
        current = user_doc.to_dict().get("credits", 0)
        new_balance = current + reserved
        users_ref().document(user_id).update({"credits": new_balance})

        transactions_ref().add({
            "user_id": user_id,
            "type": "REFUND",
            "amount": reserved,
            "balance_after": new_balance,
            "description": f"คืนเครดิตโดยแอดมิน (force refund) — {reserved} เครดิต",
            "created_at": now,
        })

    # Update job
    jobs_ref().document(job_id).update({
        "status": "REFUNDED",
        "refund_amount": reserved,
        "completed_at": now,
    })

    audit_logs_ref().add({
        "event_type": "ADMIN_FORCE_REFUND",
        "user_id": admin["user_id"],
        "details": {"job_id": job_id, "target_uid": user_id, "refunded": reserved},
        "severity": "WARNING",
        "created_at": now,
    })
    return {"message": f"คืนเครดิต {reserved} เรียบร้อย", "refunded": reserved}


@router.post("/cleanup-jobs")
async def admin_cleanup_jobs(admin: dict = Depends(require_admin)):
    """Bulk cleanup all stuck RESERVED jobs — refund credits and mark EXPIRED."""
    now = datetime.now(timezone.utc)
    cutoff = now - timedelta(hours=settings.JOB_EXPIRE_HOURS)
    count = 0
    total_refunded = 0

    for doc in jobs_ref().where("status", "==", "RESERVED").stream():
        d = doc.to_dict()
        expires = d.get("expires_at")
        created = d.get("created_at")

        # Skip if not yet expired
        if expires and expires >= now:
            continue
        if not expires and created and created >= cutoff:
            continue

        job_id = doc.id
        user_id = d.get("user_id", "")
        reserved = d.get("reserved_credits", 0)

        if user_id and reserved > 0:
            users_ref().document(user_id).update({
                "credits": firestore.Increment(reserved),
            })
            user_doc = users_ref().document(user_id).get()
            balance = user_doc.to_dict().get("credits", 0) if user_doc.exists else 0

            transactions_ref().add({
                "user_id": user_id,
                "type": "REFUND",
                "amount": reserved,
                "balance_after": balance,
                "reference_id": d.get("job_token", ""),
                "description": f"คืนเครดิตอัตโนมัติ (admin cleanup) — {reserved} เครดิต",
                "created_at": now,
            })
            total_refunded += reserved

        jobs_ref().document(job_id).update({
            "status": "EXPIRED",
            "refund_amount": reserved,
            "completed_at": now,
        })

        audit_logs_ref().add({
            "event_type": "ADMIN_BULK_CLEANUP",
            "user_id": admin["user_id"],
            "details": {"job_id": job_id, "target_uid": user_id, "refunded": reserved},
            "severity": "INFO",
            "created_at": now,
        })
        count += 1

    return {"message": f"คืนเครดิต {count} งาน รวม {total_refunded} เครดิต", "cleaned": count, "total_refunded": total_refunded}


# ══════════════════════════════════════════════════
# 5. Finance
# ══════════════════════════════════════════════════

@router.get("/finance/daily")
async def finance_daily(
    date_from: str = Query(alias="from", default=""),
    date_to: str = Query(alias="to", default=""),
    admin: dict = Depends(require_admin),
):
    """Get daily finance breakdown."""
    now_th = datetime.now(TH_TZ)

    if date_from:
        start = datetime.strptime(date_from, "%Y-%m-%d").replace(tzinfo=TH_TZ)
    else:
        start = (now_th - timedelta(days=30)).replace(hour=0, minute=0, second=0, microsecond=0)

    if date_to:
        end = datetime.strptime(date_to, "%Y-%m-%d").replace(tzinfo=TH_TZ) + timedelta(days=1)
    else:
        end = now_th.replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(days=1)

    start_utc = start.astimezone(timezone.utc)
    end_utc = end.astimezone(timezone.utc)

    # Fetch data
    all_slips = [s.to_dict() for s in slips_ref().where("status", "==", "VERIFIED").stream()]
    all_jobs = [j.to_dict() for j in jobs_ref().where("status", "==", "COMPLETED").stream()]
    all_users = [u.to_dict() for u in users_ref().stream()]

    cfg_doc = system_config_ref().document("app_settings").get()
    cfg = cfg_doc.to_dict() if cfg_doc.exists else {}
    exchange_rate = cfg.get("exchange_rate", settings.EXCHANGE_RATE)

    days_data = []
    total_topup = 0.0
    total_recognized = 0.0
    total_new_users = 0
    total_jobs = 0
    total_files = 0

    current = start
    while current < end:
        day_end = current + timedelta(days=1)
        day_str = current.strftime("%Y-%m-%d")
        d_start_utc = current.astimezone(timezone.utc)
        d_end_utc = day_end.astimezone(timezone.utc)

        topup_thb = sum(
            float(s.get("amount_detected", 0) or 0)
            for s in all_slips
            if s.get("verified_at") and d_start_utc <= s["verified_at"].astimezone(timezone.utc) < d_end_utc
        )
        topup_count = sum(
            1 for s in all_slips
            if s.get("verified_at") and d_start_utc <= s["verified_at"].astimezone(timezone.utc) < d_end_utc
        )
        recognized_cr = sum(
            j.get("actual_usage", 0)
            for j in all_jobs
            if j.get("completed_at") and d_start_utc <= j["completed_at"].astimezone(timezone.utc) < d_end_utc
        )
        recognized_thb = recognized_cr / exchange_rate if exchange_rate else 0

        new_users = sum(
            1 for u in all_users
            if u.get("created_at") and d_start_utc <= u["created_at"].astimezone(timezone.utc) < d_end_utc
        )
        active_users = sum(
            1 for u in all_users
            if u.get("last_active") and d_start_utc <= u["last_active"].astimezone(timezone.utc) < d_end_utc
        )
        jobs_count = sum(
            1 for j in all_jobs
            if j.get("completed_at") and d_start_utc <= j["completed_at"].astimezone(timezone.utc) < d_end_utc
        )
        files_processed = sum(
            j.get("success_count", 0)
            for j in all_jobs
            if j.get("completed_at") and d_start_utc <= j["completed_at"].astimezone(timezone.utc) < d_end_utc
        )

        days_data.append({
            "date": day_str,
            "topup_thb": round(topup_thb, 2),
            "topup_count": topup_count,
            "recognized_thb": round(recognized_thb, 2),
            "recognized_credits": recognized_cr,
            "new_users": new_users,
            "active_users": active_users,
            "jobs_count": jobs_count,
            "files_processed": files_processed,
        })

        total_topup += topup_thb
        total_recognized += recognized_thb
        total_new_users += new_users
        total_jobs += jobs_count
        total_files += files_processed
        current = day_end

    return {
        "days": days_data,
        "summary": {
            "total_topup_thb": round(total_topup, 2),
            "total_recognized_thb": round(total_recognized, 2),
            "total_new_users": total_new_users,
            "total_jobs": total_jobs,
            "total_files": total_files,
        },
    }


@router.get("/finance/monthly")
async def finance_monthly(
    year: int = Query(default=0),
    admin: dict = Depends(require_admin),
):
    """Get monthly finance summary for a year."""
    if year == 0:
        year = datetime.now(TH_TZ).year

    all_slips = [s.to_dict() for s in slips_ref().where("status", "==", "VERIFIED").stream()]
    all_jobs = [j.to_dict() for j in jobs_ref().where("status", "==", "COMPLETED").stream()]
    all_users = [u.to_dict() for u in users_ref().stream()]

    cfg_doc = system_config_ref().document("app_settings").get()
    cfg = cfg_doc.to_dict() if cfg_doc.exists else {}
    exchange_rate = cfg.get("exchange_rate", settings.EXCHANGE_RATE)

    months_data = []
    ytd_topup = 0.0
    ytd_recognized = 0.0
    ytd_new_users = 0
    ytd_jobs = 0

    for month in range(1, 13):
        m_start = datetime(year, month, 1, tzinfo=TH_TZ)
        if month == 12:
            m_end = datetime(year + 1, 1, 1, tzinfo=TH_TZ)
        else:
            m_end = datetime(year, month + 1, 1, tzinfo=TH_TZ)

        m_start_utc = m_start.astimezone(timezone.utc)
        m_end_utc = m_end.astimezone(timezone.utc)

        topup_thb = sum(
            float(s.get("amount_detected", 0) or 0)
            for s in all_slips
            if s.get("verified_at") and m_start_utc <= s["verified_at"].astimezone(timezone.utc) < m_end_utc
        )
        recognized_cr = sum(
            j.get("actual_usage", 0)
            for j in all_jobs
            if j.get("completed_at") and m_start_utc <= j["completed_at"].astimezone(timezone.utc) < m_end_utc
        )
        recognized_thb = recognized_cr / exchange_rate if exchange_rate else 0
        deferred = topup_thb - recognized_thb

        new_users = sum(
            1 for u in all_users
            if u.get("created_at") and m_start_utc <= u["created_at"].astimezone(timezone.utc) < m_end_utc
        )
        active_users = sum(
            1 for u in all_users
            if u.get("last_active") and m_start_utc <= u["last_active"].astimezone(timezone.utc) < m_end_utc
        )
        jobs_count = sum(
            1 for j in all_jobs
            if j.get("completed_at") and m_start_utc <= j["completed_at"].astimezone(timezone.utc) < m_end_utc
        )

        avg_per_user = round(recognized_thb / active_users, 2) if active_users else 0

        months_data.append({
            "month": f"{year}-{month:02d}",
            "topup_thb": round(topup_thb, 2),
            "recognized_thb": round(recognized_thb, 2),
            "deferred_revenue": round(deferred, 2),
            "new_users": new_users,
            "active_users": active_users,
            "jobs_count": jobs_count,
            "avg_revenue_per_user": avg_per_user,
        })

        ytd_topup += topup_thb
        ytd_recognized += recognized_thb
        ytd_new_users += new_users
        ytd_jobs += jobs_count

    return {
        "months": months_data,
        "ytd": {
            "total_topup_thb": round(ytd_topup, 2),
            "total_recognized_thb": round(ytd_recognized, 2),
            "total_deferred": round(ytd_topup - ytd_recognized, 2),
            "tax_base_estimate": round(ytd_recognized, 2),
        },
    }


@router.get("/finance/export")
async def finance_export(
    date_from: str = Query(alias="from", default=""),
    date_to: str = Query(alias="to", default=""),
    format: str = Query(default="xlsx"),
    admin: dict = Depends(require_admin),
):
    """Export finance data as Excel or PDF."""
    # Get daily data
    daily = await finance_daily(date_from=date_from, date_to=date_to, admin=admin)

    if format == "xlsx":
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            raise HTTPException(status_code=500, detail="openpyxl not installed")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "รายงานการเงิน"

        # Header
        headers = ["วันที่", "รายรับ (บาท)", "จำนวนเติม", "รายได้รับรู้ (บาท)", "เครดิตใช้", "ผู้ใช้ใหม่", "ผู้ใช้งาน", "จำนวนงาน", "ไฟล์สำเร็จ"]
        header_fill = PatternFill(start_color="0C1021", end_color="0C1021", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_idx, day in enumerate(daily["days"], 2):
            ws.cell(row=row_idx, column=1, value=day["date"])
            ws.cell(row=row_idx, column=2, value=day["topup_thb"])
            ws.cell(row=row_idx, column=3, value=day["topup_count"])
            ws.cell(row=row_idx, column=4, value=day["recognized_thb"])
            ws.cell(row=row_idx, column=5, value=day["recognized_credits"])
            ws.cell(row=row_idx, column=6, value=day["new_users"])
            ws.cell(row=row_idx, column=7, value=day["active_users"])
            ws.cell(row=row_idx, column=8, value=day["jobs_count"])
            ws.cell(row=row_idx, column=9, value=day["files_processed"])

        # Summary row
        summary_row = len(daily["days"]) + 3
        ws.cell(row=summary_row, column=1, value="รวม").font = Font(bold=True)
        s = daily["summary"]
        ws.cell(row=summary_row, column=2, value=s["total_topup_thb"]).font = Font(bold=True)
        ws.cell(row=summary_row, column=4, value=s["total_recognized_thb"]).font = Font(bold=True)
        ws.cell(row=summary_row, column=6, value=s["total_new_users"]).font = Font(bold=True)
        ws.cell(row=summary_row, column=8, value=s["total_jobs"]).font = Font(bold=True)
        ws.cell(row=summary_row, column=9, value=s["total_files"]).font = Font(bold=True)

        # Auto-width
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 12)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"bigeye_finance_{date_from or 'all'}_{date_to or 'now'}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    raise HTTPException(status_code=400, detail=f"Unsupported format: {format}")


# ══════════════════════════════════════════════════
# 6. System Config
# ══════════════════════════════════════════════════

@router.get("/config")
async def get_config(admin: dict = Depends(require_admin)):
    """Get all system configuration."""
    doc = system_config_ref().document("app_settings").get()
    if not doc.exists:
        return {}
    cfg = doc.to_dict()
    # Remove sensitive fields
    cfg.pop("_internal", None)
    return cfg


class VersionConfigRequest(BaseModel):
    app_latest_version: str = ""
    force_update_below: str = ""
    app_download_url: str = ""
    app_update_notes: str = ""


class RatesConfigRequest(BaseModel):
    credit_rates: dict = {}
    exchange_rate: int = 0


class BankConfigRequest(BaseModel):
    bank_name: str = ""
    account_number: str = ""
    account_name: str = ""


class ProcessingConfigRequest(BaseModel):
    context_cache_threshold: int = 20
    max_concurrent_images: int = 5
    max_concurrent_videos: int = 2


class MaintenanceConfigRequest(BaseModel):
    maintenance_mode: bool = False
    maintenance_message: str = ""


class PromptConfigRequest(BaseModel):
    content: str = ""


class BlacklistConfigRequest(BaseModel):
    terms: list[str] = []


@router.put("/config/version")
async def update_version_config(req: VersionConfigRequest, admin: dict = Depends(require_admin)):
    updates = {k: v for k, v in req.model_dump().items() if v}
    if updates:
        system_config_ref().document("app_settings").update(updates)
        audit_logs_ref().add({
            "event_type": "ADMIN_UPDATE_CONFIG", "user_id": admin["user_id"],
            "details": {"section": "version", "changes": updates},
            "severity": "INFO", "created_at": datetime.now(timezone.utc),
        })
    return {"message": "อัปเดตเวอร์ชันเรียบร้อย"}


@router.put("/config/rates")
async def update_rates_config(req: RatesConfigRequest, admin: dict = Depends(require_admin)):
    updates = {}
    if req.credit_rates:
        updates["credit_rates"] = req.credit_rates
    if req.exchange_rate > 0:
        updates["exchange_rate"] = req.exchange_rate
    if updates:
        system_config_ref().document("app_settings").update(updates)
        audit_logs_ref().add({
            "event_type": "ADMIN_UPDATE_CONFIG", "user_id": admin["user_id"],
            "details": {"section": "rates", "changes": updates},
            "severity": "WARNING", "created_at": datetime.now(timezone.utc),
        })
    return {"message": "อัปเดตอัตราเครดิตเรียบร้อย"}


@router.put("/config/bank")
async def update_bank_config(req: BankConfigRequest, admin: dict = Depends(require_admin)):
    system_config_ref().document("app_settings").update({
        "bank_info": req.model_dump(),
    })
    audit_logs_ref().add({
        "event_type": "ADMIN_UPDATE_CONFIG", "user_id": admin["user_id"],
        "details": {"section": "bank"},
        "severity": "INFO", "created_at": datetime.now(timezone.utc),
    })
    return {"message": "อัปเดตข้อมูลธนาคารเรียบร้อย"}


@router.put("/config/processing")
async def update_processing_config(req: ProcessingConfigRequest, admin: dict = Depends(require_admin)):
    system_config_ref().document("app_settings").update(req.model_dump())
    audit_logs_ref().add({
        "event_type": "ADMIN_UPDATE_CONFIG", "user_id": admin["user_id"],
        "details": {"section": "processing", "changes": req.model_dump()},
        "severity": "INFO", "created_at": datetime.now(timezone.utc),
    })
    return {"message": "อัปเดตการประมวลผลเรียบร้อย"}


@router.put("/config/maintenance")
async def update_maintenance_config(req: MaintenanceConfigRequest, admin: dict = Depends(require_admin)):
    system_config_ref().document("app_settings").update(req.model_dump())
    audit_logs_ref().add({
        "event_type": "ADMIN_UPDATE_CONFIG", "user_id": admin["user_id"],
        "details": {"section": "maintenance", "changes": req.model_dump()},
        "severity": "WARNING" if req.maintenance_mode else "INFO",
        "created_at": datetime.now(timezone.utc),
    })
    return {"message": "อัปเดตโหมดปิดปรับปรุงเรียบร้อย"}


@router.put("/config/prompts/{key}")
async def update_prompt(key: str, req: PromptConfigRequest, admin: dict = Depends(require_admin)):
    if key not in ("istock", "hybrid", "single"):
        raise HTTPException(status_code=400, detail=f"Invalid prompt key: {key}")
    system_config_ref().document("app_settings").update({f"prompts.{key}": req.content})
    audit_logs_ref().add({
        "event_type": "ADMIN_UPDATE_PROMPT", "user_id": admin["user_id"],
        "details": {"prompt_key": key, "length": len(req.content)},
        "severity": "WARNING", "created_at": datetime.now(timezone.utc),
    })
    return {"message": f"อัปเดตพรอมต์ '{key}' เรียบร้อย"}


@router.put("/config/blacklist")
async def update_blacklist(req: BlacklistConfigRequest, admin: dict = Depends(require_admin)):
    system_config_ref().document("app_settings").update({"blacklist": req.terms})
    audit_logs_ref().add({
        "event_type": "ADMIN_UPDATE_BLACKLIST", "user_id": admin["user_id"],
        "details": {"count": len(req.terms)},
        "severity": "INFO", "created_at": datetime.now(timezone.utc),
    })
    return {"message": f"อัปเดตคำต้องห้าม ({len(req.terms)} คำ) เรียบร้อย"}


class DictionaryConfigRequest(BaseModel):
    words: list[str] = []


@router.get("/config/dictionary")
async def get_dictionary(admin: dict = Depends(require_admin)):
    """Get keyword dictionary from system config."""
    doc = system_config_ref().document("app_settings").get()
    if not doc.exists:
        return {"words": []}
    cfg = doc.to_dict()
    raw = cfg.get("dictionary", "")
    if isinstance(raw, str):
        words = [w.strip() for w in raw.split("\n") if w.strip()]
    elif isinstance(raw, list):
        words = raw
    else:
        words = []
    return {"words": words}


@router.put("/config/dictionary")
async def update_dictionary(req: DictionaryConfigRequest, admin: dict = Depends(require_admin)):
    """Update keyword dictionary. Stored as newline-separated string in Firestore."""
    text = "\n".join(req.words)
    system_config_ref().document("app_settings").update({"dictionary": text})
    audit_logs_ref().add({
        "event_type": "ADMIN_UPDATE_DICTIONARY", "user_id": admin["user_id"],
        "details": {"count": len(req.words)},
        "severity": "INFO", "created_at": datetime.now(timezone.utc),
    })
    return {"message": f"อัปเดตพจนานุกรม ({len(req.words)} คำ) เรียบร้อย"}


# ══════════════════════════════════════════════════
# 7. Audit Logs
# ══════════════════════════════════════════════════

@router.get("/audit-logs")
async def list_audit_logs(
    severity: str = Query(default="", max_length=20),
    days: int = Query(default=7, ge=1, le=90),
    search: str = Query(default="", max_length=200),
    page: int = Query(default=1, ge=1),
    limit: int = Query(default=100, ge=1, le=500),
    admin: dict = Depends(require_admin),
):
    """List audit log entries."""
    cutoff = datetime.now(timezone.utc) - timedelta(days=days)

    docs = list(audit_logs_ref().stream())
    items = []
    for doc in docs:
        d = doc.to_dict()
        created = d.get("created_at")
        if created and created.astimezone(timezone.utc) < cutoff:
            continue
        if severity and d.get("severity", "").upper() != severity.upper():
            continue
        if search:
            q = search.lower()
            searchable = f"{d.get('event_type', '')} {d.get('user_id', '')} {str(d.get('details', ''))}".lower()
            if q not in searchable:
                continue
        d["id"] = doc.id
        items.append(d)

    items.sort(
        key=lambda l: l.get("created_at") or datetime.min.replace(tzinfo=timezone.utc),
        reverse=True,
    )

    page_items, total = _paginate(items, page, limit)

    return {
        "logs": [
            {
                "id": l["id"],
                "event_type": l.get("event_type", ""),
                "user_id": l.get("user_id", ""),
                "severity": l.get("severity", "INFO"),
                "details": l.get("details", {}),
                "created_at": _ts_to_str(l.get("created_at")),
            }
            for l in page_items
        ],
        "total": total,
        "page": page,
        "pages": math.ceil(total / limit) if total else 1,
    }
